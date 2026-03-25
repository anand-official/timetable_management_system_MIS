#!/usr/bin/env python3
"""
Timetable Excel Export Script
Generates Excel files with timetables for classes and teachers
"""

import json
import sys
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_class_timetable_excel(data, output_path):
    """Generate Excel with class timetables"""
    wb = Workbook()
    
    # Styles
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Times New Roman', size=16, bold=True)
    subtitle_font = Font(name='Times New Roman', size=12)
    cell_font = Font(name='Times New Roman', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    sections = data.get('sections', [])
    days = data.get('days', ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
    periods = data.get('periods', [])
    timetable = data.get('timetable', [])
    
    # Group timetable by section
    section_timetables = {}
    for slot in timetable:
        section_name = slot.get('section', 'Unknown')
        if section_name not in section_timetables:
            section_timetables[section_name] = {}
        key = f"{slot.get('day', '')}-{slot.get('period', 0)}"
        section_timetables[section_name][key] = slot
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for section in sections:
        section_name = section.get('name', 'Unknown')
        
        # Create sheet for section
        ws = wb.create_sheet(title=section_name[:31])  # Excel sheet name limit
        
        # Title
        ws['B2'] = f"Modern Indian School - {section_name}"
        ws['B2'].font = title_font
        ws.merge_cells('B2:H2')
        
        ws['B3'] = f"Academic Year 2025-26 | Generated: {datetime.now().strftime('%Y-%m-%d')}"
        ws['B3'].font = subtitle_font
        ws.merge_cells('B3:H3')
        
        if section.get('classTeacher'):
            ws['B4'] = f"Class Teacher: {section['classTeacher']}"
            ws['B4'].font = subtitle_font
        
        # Header row
        row = 6
        ws.cell(row=row, column=2, value='Period').font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        
        for col_idx, day in enumerate(days, start=3):
            cell = ws.cell(row=row, column=col_idx, value=day)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Period rows
        section_slots = section_timetables.get(section_name, {})
        
        for period_idx, period_info in enumerate(periods):
            period_num = period_info.get('period', 0)
            time_str = f"{period_info.get('start', '')}-{period_info.get('end', '')}"
            
            row = 7 + period_idx
            
            # Period column
            period_cell = ws.cell(row=row, column=2, value=f"P{period_num}\n{time_str}")
            period_cell.font = Font(name='Times New Roman', size=9, bold=True)
            period_cell.alignment = center_align
            period_cell.border = thin_border
            if period_idx % 2 == 1:
                period_cell.fill = alt_fill
            
            # Day columns
            for col_idx, day in enumerate(days, start=3):
                key = f"{day}-{period_num}"
                slot = section_slots.get(key, {})
                
                if slot:
                    subject = slot.get('subject', '')
                    teacher = slot.get('teacherAbbr', '')
                    cell_value = f"{subject}\n({teacher})"
                else:
                    cell_value = '-'
                
                cell = ws.cell(row=row, column=col_idx, value=cell_value)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if period_idx % 2 == 1:
                    cell.fill = alt_fill
        
        # Set column widths
        ws.column_dimensions['B'].width = 12
        for col in range(3, 3 + len(days)):
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        # Set row heights
        for r in range(7, 7 + len(periods)):
            ws.row_dimensions[r].height = 35
    
    wb.save(output_path)
    print(f"Excel generated: {output_path}")


def create_teacher_timetable_excel(data, output_path):
    """Generate Excel with teacher timetables"""
    wb = Workbook()
    
    # Styles
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Times New Roman', size=16, bold=True)
    subtitle_font = Font(name='Times New Roman', size=12)
    cell_font = Font(name='Times New Roman', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    days = data.get('days', ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
    periods = data.get('periods', [])
    timetable = data.get('timetable', [])
    teachers = data.get('teachers', [])
    
    # Group timetable by teacher
    teacher_timetables = {}
    for slot in timetable:
        teacher_name = slot.get('teacher', 'Unknown')
        if teacher_name not in teacher_timetables:
            teacher_timetables[teacher_name] = {}
        key = f"{slot.get('day', '')}-{slot.get('period', 0)}"
        teacher_timetables[teacher_name][key] = slot
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for teacher in teachers:
        teacher_name = teacher.get('name', 'Unknown')
        abbr = teacher.get('abbreviation', '')
        dept = teacher.get('department', '')
        
        # Create sheet name (abbreviated if too long)
        sheet_name = abbr if abbr else teacher_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Title
        ws['B2'] = f"Teacher Schedule: {teacher_name}"
        ws['B2'].font = title_font
        ws.merge_cells('B2:H2')
        
        ws['B3'] = f"Department: {dept} | Target: {teacher.get('targetWorkload', 0)} | Current: {teacher.get('currentWorkload', 0)}"
        ws['B3'].font = subtitle_font
        ws.merge_cells('B3:H3')
        
        # Header row
        row = 5
        ws.cell(row=row, column=2, value='Period').font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        
        for col_idx, day in enumerate(days, start=3):
            cell = ws.cell(row=row, column=col_idx, value=day)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Period rows
        teacher_slots = teacher_timetables.get(teacher_name, {})
        
        for period_idx, period_info in enumerate(periods):
            period_num = period_info.get('period', 0)
            time_str = f"{period_info.get('start', '')}-{period_info.get('end', '')}"
            
            row = 6 + period_idx
            
            # Period column
            period_cell = ws.cell(row=row, column=2, value=f"P{period_num}\n{time_str}")
            period_cell.font = Font(name='Times New Roman', size=9, bold=True)
            period_cell.alignment = center_align
            period_cell.border = thin_border
            if period_idx % 2 == 1:
                period_cell.fill = alt_fill
            
            # Day columns
            for col_idx, day in enumerate(days, start=3):
                key = f"{day}-{period_num}"
                slot = teacher_slots.get(key, {})
                
                if slot:
                    section = slot.get('section', '')
                    subject = slot.get('subject', '')
                    cell_value = f"{section}\n({subject})"
                else:
                    cell_value = '-'
                
                cell = ws.cell(row=row, column=col_idx, value=cell_value)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                if period_idx % 2 == 1:
                    cell.fill = alt_fill
        
        # Set column widths
        ws.column_dimensions['B'].width = 12
        for col in range(3, 3 + len(days)):
            ws.column_dimensions[get_column_letter(col)].width = 16
        
        # Set row heights
        for r in range(6, 6 + len(periods)):
            ws.row_dimensions[r].height = 35
    
    wb.save(output_path)
    print(f"Teacher Excel generated: {output_path}")


def create_workload_excel(data, output_path):
    """Generate Excel with teacher workload summary"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Workload Summary"
    
    # Styles
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Times New Roman', size=16, bold=True)
    cell_font = Font(name='Times New Roman', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ok_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    under_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    over_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    # Title
    ws['B2'] = "Modern Indian School - Teacher Workload Summary"
    ws['B2'].font = title_font
    ws.merge_cells('B2:G2')
    
    ws['B3'] = f"Academic Year 2025-26 | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    
    teachers = data.get('teachers', [])
    
    # Headers
    headers = ['Teacher', 'Abbreviation', 'Department', 'Target', 'Current', 'Difference', 'Status']
    row = 5
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data rows
    for idx, teacher in enumerate(teachers):
        row = 6 + idx
        target = teacher.get('targetWorkload', 0)
        current = teacher.get('currentWorkload', 0)
        diff = current - target
        
        if abs(diff) <= 2:
            status = 'OK'
            fill = ok_fill
        elif diff < 0:
            status = 'Under'
            fill = under_fill
        else:
            status = 'Over'
            fill = over_fill
        
        values = [
            teacher.get('name', ''),
            teacher.get('abbreviation', ''),
            teacher.get('department', ''),
            target,
            current,
            diff,
            status
        ]
        
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            if col_idx == 8:  # Status column
                cell.fill = fill
    
    # Set column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    wb.save(output_path)
    print(f"Workload Excel generated: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_excel.py <data_file.json> <output_file.xlsx> [type]")
        print("  type: 'class', 'teacher', or 'workload' (default: 'class')")
        sys.exit(1)
    
    data_file = sys.argv[1]
    output_file = sys.argv[2]
    export_type = sys.argv[3] if len(sys.argv) > 3 else 'class'
    
    with open(data_file, 'r') as f:
        data = json.load(f)
    
    if export_type == 'teacher':
        create_teacher_timetable_excel(data, output_file)
    elif export_type == 'workload':
        create_workload_excel(data, output_file)
    else:
        create_class_timetable_excel(data, output_file)
